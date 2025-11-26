import reflex as rx
from typing import Any
from datetime import datetime
from pathlib import Path
import uuid
import logging
from pydantic import BaseModel
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


class PackageDimension(BaseModel):
    id: str
    bultos: int = 1
    largo: float = 0.0  # Length in inches
    ancho: float = 0.0  # Width in inches
    alto: float = 0.0  # Height in inches
    pounds: float = 0.0  # Weight in pounds
    cubic_feet: float = 0.0  # Volume (calculated)
    pt: float = 0.0  # Chargeable weight
    referencia: str = ""


class WarehouseReceiptState(rx.State):
    """State for the Warehouse Receipt document."""

    is_loading: bool = False

    # Header information
    receipt_number: str = ""
    warehouse_location: str = ""
    receipt_date: str = ""

    # Company information
    company_name: str = "Nosglobal Logistic"
    company_logo_url: str = "/nosglobal-logo.png"

    # Package summary (calculated from dimensions)
    peso_bruto: float = 0.0  # Gross weight in pounds
    volumen: float = 0.0  # Volume in cubic feet
    peso_tasable: float = 0.0  # Chargeable weight in pounds

    # Details section
    oficina: str = ""
    remitente: str = ""
    referencia: str = ""
    destinatario: str = ""
    no_pedido: str = ""
    entregado_por: str = ""
    tracking_number: str = ""
    factura: str = ""
    descripcion: str = ""

    # Package dimensions table
    dimensions: list[PackageDimension] = []

    # Legal disclaimer
    legal_disclaimer: str = (
        "Nuestra empresa no se hace responsable por pérdida o daños totales y/o parciales de mercancía "
        "que NO SE ENCUENTRE ASEGURADA. El seguro únicamente aplicará bajo previa inspección de "
        "los artículos y aprobación de los mismos. Igualmente, Nosglobal Logistic no se hace responsable de "
        "paquetes perdidos en tránsito desde su proveedor hasta nuestros almacenes, ni de paquetes que "
        "no contengan el servicio de firma requerida. Por tal razón, recomendamos que sus envíos sean "
        "manejados por empresas que puedan proveerle un número de rastreo (tracking) para de este "
        "modo tener un mayor control de su mercancía. Les recomendamos los pesos promedio por caja "
        "es de un máximo de 90 Lbs. Cajas que sobrepase los pesos permitidos, La Compañía no se hace "
        "responsable por daños en el manejo de su carga. Los equipos electrónicos como Televisores se "
        "reciben solamente como mercancía general."
    )

    @rx.event
    def on_load(self):
        """Initialize date on client load to avoid hydration mismatch."""
        if not self.receipt_date:
            self.receipt_date = datetime.now().strftime("%Y-%m-%d")

    @rx.var
    def total_bultos(self) -> int:
        return sum([d.bultos for d in self.dimensions])

    @rx.var
    def calculated_peso_bruto(self) -> float:
        return sum([d.pounds for d in self.dimensions])

    @rx.var
    def calculated_volumen(self) -> float:
        return sum([d.cubic_feet for d in self.dimensions])

    @rx.event
    def set_field(self, field: str, value: str):
        if hasattr(self, field):
            if field in ["peso_bruto", "volumen", "peso_tasable"]:
                try:
                    setattr(self, field, float(value))
                except ValueError as e:
                    logging.exception(f"Error converting value to float: {e}")
                    setattr(self, field, 0.0)
            else:
                setattr(self, field, value)

    @rx.event
    def add_dimension(self):
        self.dimensions.append(
            PackageDimension(
                id=str(uuid.uuid4()),
                bultos=1,
                largo=0.0,
                ancho=0.0,
                alto=0.0,
                pounds=0.0,
                cubic_feet=0.0,
                pt=0.0,
                referencia="",
            )
        )

    @rx.event
    def remove_dimension(self, idx: int):
        if 0 <= idx < len(self.dimensions):
            self.dimensions.pop(idx)
            # Create a new list to ensure Reflex detects the change correctly
            self.dimensions = list(self.dimensions)

    @rx.event
    def update_dimension(self, idx: int, field: str, value: Any):
        if 0 <= idx < len(self.dimensions):
            dimension = self.dimensions[idx]
            if field in ["bultos", "largo", "ancho", "alto", "pounds", "pt"]:
                try:
                    if value == "" or value is None:
                        val = 0.0
                    else:
                        val = float(value)
                    
                    if field == "bultos":
                        dimension.bultos = int(val)
                    elif field == "largo":
                        dimension.largo = val
                    elif field == "ancho":
                        dimension.ancho = val
                    elif field == "alto":
                        dimension.alto = val
                    elif field == "pounds":
                        dimension.pounds = val
                    elif field == "pt":
                        dimension.pt = val

                    # Calculate cubic feet automatically
                    if dimension.largo > 0 and dimension.ancho > 0 and dimension.alto > 0:
                        dimension.cubic_feet = (
                            dimension.largo * dimension.ancho * dimension.alto
                        ) / 1728  # Convert cubic inches to cubic feet
                    else:
                        dimension.cubic_feet = 0.0
                        
                except ValueError as e:
                    logging.warning(f"Invalid value for field {field}: {value}")
                    # Don't update state on invalid input to prevent UI glitches,
                    # but allow the user to correct it.
            elif field == "referencia":
                dimension.referencia = str(value)

            # Create a new list to ensure Reflex detects the change correctly
            # This is critical to prevent "NotFoundError: removeChild" errors
            self.dimensions = list(self.dimensions)

    @rx.event
    async def export_pdf(self):
        self.is_loading = True
        filename = f"ReciboAlmacen_{self.receipt_number}_{uuid.uuid4().hex[:6]}.pdf"
        upload_dir = Path(".web/public")
        upload_dir.mkdir(parents=True, exist_ok=True)
        file_path = upload_dir / filename
        try:
            doc = SimpleDocTemplate(
                str(file_path),
                pagesize=letter,
                rightMargin=40,
                leftMargin=40,
                topMargin=40,
                bottomMargin=40,
            )
            elements = []
            styles = getSampleStyleSheet()

            # Header with logo support
            logo_path = Path(".web/public") / self.company_logo_url.lstrip("/")
            left_content = []

            # Add logo if it exists
            if logo_path.exists():
                try:
                    logo = Image(str(logo_path), width=0.8*inch, height=0.8*inch)
                    left_content.append([logo])
                except Exception as e:
                    logging.warning(f"Could not load logo: {e}")

            # Add company name
            left_content.append([Paragraph(f"<b>{self.company_name}</b>", styles["Normal"])])

            # Create nested table for left column if logo exists
            if len(left_content) > 1:
                left_table = Table(left_content, colWidths=[1.5 * inch])
                left_table.setStyle(TableStyle([
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 0),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ]))
                left_cell = left_table
            else:
                left_cell = Paragraph(f"<b>{self.company_name}</b>", styles["Normal"])

            header_data = [
                [
                    left_cell,
                    Paragraph(
                        f"<font size=18><b>RECIBO DE ALMACÉN</b></font><br/><br/>"
                        f"<font size=16><b>{self.receipt_number}</b></font><br/><br/>"
                        f"<font size=12><b>{self.warehouse_location}</b></font>",
                        styles["Normal"],
                    ),
                ]
            ]
            t_header = Table(header_data, colWidths=[4 * inch, 3 * inch])
            t_header.setStyle(
                TableStyle(
                    [
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                        ("ALIGN", (1, 0), (1, 0), "RIGHT"),
                    ]
                )
            )
            elements.append(t_header)
            elements.append(Spacer(1, 30))

            # Details section
            details_data = [
                ["Fecha", self.receipt_date, "Oficina", self.oficina],
                ["Remitente", self.remitente, "Referencia", self.referencia],
                ["Destinatario", self.destinatario, "No. Pedido", self.no_pedido],
                ["Entregado por", self.entregado_por, "Factura", self.factura],
                ["Tracking", self.tracking_number, "", ""],
                ["Descripción", self.descripcion, "", ""],
            ]
            t_details = Table(
                details_data, colWidths=[1.5 * inch, 2 * inch, 1.5 * inch, 2 * inch]
            )
            t_details.setStyle(
                TableStyle(
                    [
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                        ("BACKGROUND", (0, 0), (-1, -1), colors.Color(0.96, 0.96, 0.96)),  # Light grey background
                        ("FONTSIZE", (0, 0), (-1, -1), 9),
                        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                        ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                        ("PADDING", (0, 0), (-1, -1), 6),
                    ]
                )
            )
            elements.append(t_details)
            elements.append(Spacer(1, 30))

            # Dimensions table
            elements.append(Paragraph("<b>Dimensiones de Paquetes</b>", styles["Heading4"]))
            elements.append(Spacer(1, 10))

            dim_headers = [
                "Bultos",
                "Largo",
                "Ancho",
                "Alto",
                "Pounds",
                "Cubic Feet",
                "PT",
                "Referencia",
            ]
            dim_data = [dim_headers]
            for d in self.dimensions:
                dim_data.append(
                    [
                        str(d.bultos),
                        f"{d.largo:.1f}" if d.largo > 0 else "X",
                        f"{d.ancho:.1f}" if d.ancho > 0 else "X",
                        f"{d.alto:.1f}" if d.alto > 0 else "X",
                        f"{d.pounds:.1f} lbs",
                        f"{d.cubic_feet:.3f}",
                        str(d.pt) if d.pt > 0 else "",
                        d.referencia,
                    ]
                )
            t_dimensions = Table(
                dim_data,
                colWidths=[
                    0.6 * inch,
                    0.7 * inch,
                    0.7 * inch,
                    0.7 * inch,
                    0.9 * inch,
                    1 * inch,
                    0.6 * inch,
                    1.3 * inch,
                ],
            )
            t_dimensions.setStyle(
                TableStyle(
                    [
                        # Remove full grid, use only horizontal lines for row separation
                        ("LINEBELOW", (0, 0), (-1, 0), 1, colors.grey),  # Bold line under header
                        ("LINEBELOW", (0, 1), (-1, -1), 0.25, colors.Color(0.9, 0.9, 0.9)),  # Light lines under rows
                        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                        ("FONTSIZE", (0, 0), (-1, -1), 9),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                        ("PADDING", (0, 0), (-1, -1), 6),
                    ]
                )
            )
            elements.append(t_dimensions)
            elements.append(Spacer(1, 30))

            # Archive section
            elements.append(Paragraph("<b>Archivo</b>", styles["Heading4"]))
            elements.append(Spacer(1, 5))

            # Create a styled box for the archive message
            archive_data = [["No se han encontrado registros"]]
            t_archive = Table(archive_data, colWidths=[7 * inch])
            t_archive.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, -1), colors.Color(0.96, 0.96, 0.96)),
                        ("FONTSIZE", (0, 0), (-1, -1), 9),
                        ("TEXTCOLOR", (0, 0), (-1, -1), colors.Color(0.5, 0.5, 0.5)),
                        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                        ("PADDING", (0, 0), (-1, -1), 8),
                        ("BOX", (0, 0), (-1, -1), 0.5, colors.Color(0.85, 0.85, 0.85)),
                    ]
                )
            )
            elements.append(t_archive)
            elements.append(Spacer(1, 30))

            # Legal disclaimer with top border
            disclaimer_data = [[Paragraph(f"<font size=7>{self.legal_disclaimer}</font>", styles["Normal"])]]
            t_disclaimer = Table(disclaimer_data, colWidths=[7 * inch])
            t_disclaimer.setStyle(
                TableStyle(
                    [
                        ("LINEABOVE", (0, 0), (-1, 0), 1, colors.Color(0.85, 0.85, 0.85)),
                        ("TOPPADDING", (0, 0), (-1, -1), 15),
                        ("LEFTPADDING", (0, 0), (-1, -1), 0),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                    ]
                )
            )
            elements.append(t_disclaimer)

            doc.build(elements)
            self.is_loading = False
            if not file_path.exists():
                raise FileNotFoundError(f"PDF file was not created: {file_path}")

            # Read file data and pass directly to download
            with open(file_path, "rb") as f:
                pdf_data = f.read()

            # Clean up temporary file
            try:
                file_path.unlink()
            except Exception as e:
                logging.warning(f"Could not delete temporary file: {e}")

            return rx.download(data=pdf_data, filename=filename)
        except Exception as e:
            self.is_loading = False
            logging.exception(f"PDF Generation Error: {e}")
            return rx.toast.error(f"Error generating PDF: {str(e)}")

    @rx.event
    async def export_excel(self):
        self.is_loading = True
        filename = f"ReciboAlmacen_{self.receipt_number}_{uuid.uuid4().hex[:6]}.xlsx"
        upload_dir = Path(".web/public")
        upload_dir.mkdir(parents=True, exist_ok=True)
        file_path = upload_dir / filename
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Recibo de Almacen"

            title_font = Font(bold=True, size=16)
            header_font = Font(bold=True)
            gray_fill = PatternFill(
                start_color="EEEEEE", end_color="EEEEEE", fill_type="solid"
            )

            # Header
            ws["A1"] = self.company_name
            ws["A1"].font = title_font
            ws["E1"] = f"RECIBO DE ALMACÉN {self.receipt_number}"
            ws["E1"].font = title_font
            ws["E2"] = self.warehouse_location
            ws["E2"].font = header_font

            # Summary
            ws["A4"] = "Bultos"
            ws["B4"] = "Peso Bruto"
            ws["C4"] = "Volumen"
            ws["D4"] = "Peso Tasable"
            for cell in ["A4", "B4", "C4", "D4"]:
                ws[cell].font = header_font
                ws[cell].fill = gray_fill

            ws["A5"] = self.total_bultos
            ws["B5"] = f"{self.calculated_peso_bruto:.2f} pound(s)"
            ws["C5"] = f"{self.calculated_volumen:.3f} cubic feet"
            ws["D5"] = f"{self.peso_tasable:.2f} pound(s)"

            # Details
            row = 7
            details = [
                ["Fecha", self.receipt_date],
                ["Oficina", self.oficina],
                ["Remitente", self.remitente],
                ["Referencia", self.referencia],
                ["Destinatario", self.destinatario],
                ["No. Pedido", self.no_pedido],
                ["Entregado por", self.entregado_por],
                ["Tracking", self.tracking_number],
                ["Factura", self.factura],
                ["Descripción", self.descripcion],
            ]
            for label, value in details:
                ws.cell(row=row, column=1, value=label).font = header_font
                ws.cell(row=row, column=2, value=value)
                row += 1

            # Dimensions table
            row += 2
            headers = [
                "Bultos",
                "Largo",
                "Ancho",
                "Alto",
                "Pounds",
                "Cubic Feet",
                "PT",
                "Referencia",
            ]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = gray_fill

            row += 1
            for d in self.dimensions:
                ws.cell(row=row, column=1, value=d.bultos)
                ws.cell(
                    row=row, column=2, value=d.largo if d.largo > 0 else "X"
                )
                ws.cell(
                    row=row, column=3, value=d.ancho if d.ancho > 0 else "X"
                )
                ws.cell(
                    row=row, column=4, value=d.alto if d.alto > 0 else "X"
                )
                ws.cell(row=row, column=5, value=d.pounds)
                ws.cell(row=row, column=6, value=d.cubic_feet)
                ws.cell(row=row, column=7, value=d.pt)
                ws.cell(row=row, column=8, value=d.referencia)
                row += 1

            # Archive section
            row += 2
            ws.cell(row=row, column=1, value="ARCHIVO:").font = header_font
            row += 1
            ws.cell(row=row, column=1, value="No se han encontrado registros")

            # Adjust column widths
            ws.column_dimensions["A"].width = 20
            ws.column_dimensions["B"].width = 30
            ws.column_dimensions["C"].width = 15
            ws.column_dimensions["D"].width = 15
            ws.column_dimensions["E"].width = 25

            wb.save(file_path)
            self.is_loading = False

            # Read file data and pass directly to download
            with open(file_path, "rb") as f:
                excel_data = f.read()

            # Clean up temporary file
            try:
                file_path.unlink()
            except Exception as e:
                logging.warning(f"Could not delete temporary file: {e}")

            return rx.download(data=excel_data, filename=filename)
        except Exception as e:
            self.is_loading = False
            logging.exception(f"Excel Generation Error: {e}")
            return rx.toast.error(f"Error generating Excel: {str(e)}")
