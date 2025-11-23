import reflex as rx
from typing import Any
from datetime import datetime, date
from pathlib import Path
import uuid
import os
import logging
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
    Image,
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


class Transaction(rx.Base):
    id: str
    date: str
    invoice_no: str
    reference: str
    description: str
    amount: float
    paid: float


class StatementState(rx.State):
    """State for the Account Statement document."""

    is_loading: bool = False
    provider_name: str = "Nosglobal Logistic"
    provider_address: str = "Av. Principal 1000, Torre A, Piso 5"
    provider_city_state_zip: str = "Caracas, Distrito Capital 1010"
    provider_phone: str = "+58 212 123 4567"
    client_name: str = ""
    client_address: str = ""
    client_city: str = ""
    client_state: str = ""
    client_country: str = ""
    account_number: str = ""
    terms: str = ""
    statement_date: str = ""
    transactions: list[Transaction] = []

    @rx.event
    def on_load(self):
        """Initialize date on client load to avoid hydration mismatch."""
        if not self.statement_date:
            self.statement_date = datetime.now().strftime("%Y-%m-%d")

    @rx.var
    def total_due(self) -> float:
        return sum([t.amount - t.paid for t in self.transactions])

    @rx.var
    def aging_buckets(self) -> dict[str, float]:
        current = 0.0
        days_30 = 0.0
        days_60 = 0.0
        days_90 = 0.0

        # Use today's date if statement_date is empty
        if not self.statement_date:
            stmt_date = date.today()
        else:
            try:
                stmt_date = datetime.strptime(self.statement_date, "%Y-%m-%d").date()
            except ValueError:
                stmt_date = date.today()

        for t in self.transactions:
            if not t.date:
                inv_date = date.today()
            else:
                try:
                    inv_date = datetime.strptime(t.date, "%Y-%m-%d").date()
                except ValueError:
                    inv_date = date.today()
            days_diff = (stmt_date - inv_date).days
            balance = t.amount - t.paid
            if days_diff < 30:
                current += balance
            elif days_diff < 60:
                days_30 += balance
            elif days_diff < 90:
                days_60 += balance
            else:
                days_90 += balance
        return {"current": current, "30": days_30, "60": days_60, "90": days_90}

    @rx.event
    def set_field(self, field: str, value: str):
        setattr(self, field, value)

    @rx.event
    def add_transaction(self):
        self.transactions.append(
            Transaction(
                id=str(uuid.uuid4()),
                date=datetime.now().strftime("%Y-%m-%d"),
                invoice_no="",
                reference="",
                description="",
                amount=0.0,
                paid=0.0,
            )
        )

    @rx.event
    def update_transaction(self, idx: int, field: str, value: Any):
        if 0 <= idx < len(self.transactions):
            transaction = self.transactions[idx]
            if field in ["amount", "paid"]:
                try:
                    setattr(transaction, field, float(value))
                except ValueError as e:
                    logging.exception(f"Error converting value to float: {e}")
                    setattr(transaction, field, 0.0)
            else:
                setattr(transaction, field, value)
            self.transactions = self.transactions

    @rx.event
    def remove_transaction(self, idx: int):
        if 0 <= idx < len(self.transactions):
            self.transactions.pop(idx)
            self.transactions = self.transactions

    @rx.event
    async def export_pdf(self):
        self.is_loading = True
        filename = f"Statement_{self.account_number}_{uuid.uuid4().hex[:6]}.pdf"
        upload_dir = Path(".web/public")
        upload_dir.mkdir(parents=True, exist_ok=True)
        file_path = upload_dir / filename
        try:
            doc = SimpleDocTemplate(
                str(file_path),
                pagesize=letter,
                rightMargin=30,
                leftMargin=30,
                topMargin=30,
                bottomMargin=30,
            )
            elements = []
            styles = getSampleStyleSheet()
            header_data = [
                [
                    Paragraph(
                        f"<b>{self.provider_name}</b><br/>{self.provider_address}<br/>{self.provider_city_state_zip}<br/>Tel: {self.provider_phone}",
                        styles["Normal"],
                    ),
                    Paragraph(f"<b>ESTADO DE CUENTA</b>", styles["Heading1"]),
                ]
            ]
            t_header = Table(header_data, colWidths=[4 * inch, 3 * inch])
            t_header.setStyle(
                TableStyle(
                    [
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                        ("ALIGN", (1, 0), (1, 0), "RIGHT"),
                    ]
                )
            )
            elements.append(t_header)
            elements.append(Spacer(1, 20))
            info_data = [
                [
                    Paragraph(
                        f"<b>{self.client_name}</b><br/>{self.client_address}<br/>{self.client_city} {self.client_state}<br/>{self.client_country}",
                        styles["Normal"],
                    ),
                    Table(
                        [
                            ["NÚMERO DE CUENTA", self.account_number],
                            ["TÉRMINOS", self.terms],
                            ["FECHA ESTADO", self.statement_date],
                        ],
                        style=TableStyle(
                            [
                                ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                                ("BACKGROUND", (0, 0), (0, -1), colors.lightgrey),
                                ("FONTSIZE", (0, 0), (-1, -1), 8),
                                ("PADDING", (0, 0), (-1, -1), 4),
                            ]
                        ),
                    ),
                ]
            ]
            t_info = Table(info_data, colWidths=[4.5 * inch, 2.5 * inch])
            t_info.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP")]))
            elements.append(t_info)
            elements.append(Spacer(1, 20))
            elements.append(
                Paragraph(
                    f"A CONTINUACION LE MOSTRAMOS UNA LISTA DE NOTAS DE ENTREGA PENDIENTES DE PAGO A {self.statement_date}",
                    styles["Normal"],
                )
            )
            elements.append(Spacer(1, 10))
            trans_data = [
                [
                    "FECHA",
                    "NOTA DE ENTREGA",
                    "CUENTA",
                    "DESCRIPCIÓN",
                    "CANTIDAD",
                    "PAGADO",
                    "DEBIDO",
                ]
            ]
            for t in self.transactions:
                trans_data.append(
                    [
                        t.date,
                        t.invoice_no,
                        t.reference,
                        Paragraph(t.description, styles["Normal"]),
                        f"{t.amount:,.2f}",
                        f"{t.paid:,.2f}",
                        f"{t.amount - t.paid:,.2f}",
                    ]
                )
            t_trans = Table(
                trans_data,
                colWidths=[
                    1 * inch,
                    0.8 * inch,
                    1 * inch,
                    2.2 * inch,
                    0.8 * inch,
                    0.8 * inch,
                    0.9 * inch,
                ],
            )
            t_trans.setStyle(
                TableStyle(
                    [
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                        ("FONTSIZE", (0, 0), (-1, -1), 8),
                        ("ALIGN", (4, 1), (-1, -1), "RIGHT"),
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                        ("PADDING", (0, 0), (-1, -1), 4),
                    ]
                )
            )
            elements.append(t_trans)
            elements.append(Spacer(1, 20))
            aging = self.aging_buckets
            aging_data = [
                ["CURRENCY", "-30", "+30", "+60", "+90", "TOTAL DEBIDO"],
                [
                    "USD",
                    f"{aging['current']:,.2f}",
                    f"{aging['30']:,.2f}",
                    f"{aging['60']:,.2f}",
                    f"{aging['90']:,.2f}",
                    f"{self.total_due:,.2f}",
                ],
            ]
            t_aging = Table(
                aging_data,
                colWidths=[
                    1 * inch,
                    1 * inch,
                    1 * inch,
                    1 * inch,
                    1 * inch,
                    1.5 * inch,
                ],
            )
            t_aging.setStyle(
                TableStyle(
                    [
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
                        ("FONTSIZE", (0, 0), (-1, -1), 9),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ]
                )
            )
            elements.append(t_aging)
            doc.build(elements)
            self.is_loading = False
            if not file_path.exists():
                raise FileNotFoundError(f"PDF file was not created: {file_path}")

            # Read file data and pass directly to download
            with open(file_path, "rb") as f:
                pdf_data = f.read()

            # Clean up temporary file
            try:
                file_path.unlink()
            except Exception as e:
                logging.warning(f"Could not delete temporary file: {e}")

            return rx.download(data=pdf_data, filename=filename)
        except Exception as e:
            self.is_loading = False
            logging.exception(f"PDF Generation Error: {e}")
            return rx.toast.error(f"Error generating PDF: {str(e)}")

    @rx.event
    async def export_excel(self):
        self.is_loading = True
        filename = f"Statement_{self.account_number}_{uuid.uuid4().hex[:6]}.xlsx"
        upload_dir = Path(".web/public")
        upload_dir.mkdir(parents=True, exist_ok=True)
        file_path = upload_dir / filename
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Estado de Cuenta"
            ws["A1"] = self.provider_name
            ws["A2"] = self.provider_address
            ws["A3"] = self.provider_city_state_zip
            ws["E1"] = "ESTADO DE CUENTA"
            ws["E1"].font = Font(bold=True, size=14)
            ws["A6"] = "CLIENTE:"
            ws["A7"] = self.client_name
            ws["A8"] = self.client_address
            ws["A9"] = f"{self.client_city} {self.client_state}"
            ws["A10"] = self.client_country
            ws["E6"] = "NÚMERO DE CUENTA"
            ws["F6"] = self.account_number
            ws["E7"] = "TÉRMINOS"
            ws["F7"] = self.terms
            ws["E8"] = "FECHA"
            ws["F8"] = self.statement_date
            headers = [
                "FECHA",
                "NOTA DE ENTREGA",
                "CUENTA",
                "DESCRIPCIÓN",
                "CANTIDAD",
                "PAGADO",
                "DEBIDO",
            ]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=13, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(
                    start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
                )
            row = 14
            for t in self.transactions:
                ws.cell(row=row, column=1, value=t.date)
                ws.cell(row=row, column=2, value=t.invoice_no)
                ws.cell(row=row, column=3, value=t.reference)
                ws.cell(row=row, column=4, value=t.description)
                ws.cell(row=row, column=5, value=t.amount)
                ws.cell(row=row, column=6, value=t.paid)
                ws.cell(row=row, column=7, value=t.amount - t.paid)
                row += 1
            row += 2
            aging = self.aging_buckets
            ws.cell(row=row, column=1, value="AGING")
            ws.cell(row=row + 1, column=1, value="Current")
            ws.cell(row=row + 1, column=2, value=aging["current"])
            ws.cell(row=row + 1, column=3, value="30 Days")
            ws.cell(row=row + 1, column=4, value=aging["30"])
            ws.cell(row=row + 1, column=5, value="60 Days")
            ws.cell(row=row + 1, column=6, value=aging["60"])
            ws.cell(row=row + 1, column=7, value="Total Due")
            ws.cell(row=row + 1, column=8, value=self.total_due)
            wb.save(file_path)
            self.is_loading = False

            # Read file data and pass directly to download
            with open(file_path, "rb") as f:
                excel_data = f.read()

            # Clean up temporary file
            try:
                file_path.unlink()
            except Exception as e:
                logging.warning(f"Could not delete temporary file: {e}")

            return rx.download(data=excel_data, filename=filename)
        except Exception as e:
            self.is_loading = False
            logging.exception(f"Excel Generation Error: {e}")
            return rx.toast.error(f"Error generating Excel: {str(e)}")