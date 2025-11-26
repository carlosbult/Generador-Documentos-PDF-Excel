"""Quotation state management."""

import logging
import uuid
from datetime import datetime, timedelta
from pathlib import Path

import reflex as rx
from pydantic import BaseModel
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


class QuotationItem(BaseModel):
    """Model for quotation line items."""

    id: str
    description: str = ""
    quantity: int = 1
    unit_price: float = 0.0
    discount: float = 0.0
    amount: float = 0.0
    notes: str = ""


class QuotationState(rx.State):
    """State management for quotation generation."""

    is_loading: bool = False

    # Company/Header info
    company_name: str = "Nosglobal Logistic"
    company_logo_url: str = "/nosglobal-logo.png"
    company_address: str = "Miami, FL"
    company_phone: str = "+58 424-4966616"
    company_email: str = "info@nosglobal.com"

    # Quotation details
    quote_number: str = ""
    quote_date: str = ""
    valid_until: str = ""

    # Client information (REQUIRED)
    client_name: str = ""
    client_company: str = ""
    client_address: str = ""

    # Optional client details
    client_email: str = ""
    client_phone: str = ""

    # Line items
    items: list[QuotationItem] = []

    # Financial fields
    tax_rate: float = 0.0
    shipping_cost: float = 0.0
    discount_global: float = 0.0

    # Additional sections
    notes: str = ""
    terms_conditions: str = ""
    payment_terms: str = ""

    # Logo URL
    logo_url: str = "/nosglobal-logo.png"

    @rx.var
    def subtotal(self) -> float:
        """Calculate subtotal of all items."""
        return sum([item.amount for item in self.items])

    @rx.var
    def discount_total(self) -> float:
        """Calculate total discount."""
        return self.discount_global

    @rx.var
    def subtotal_after_discount(self) -> float:
        """Calculate subtotal after global discount."""
        return self.subtotal - self.discount_global

    @rx.var
    def tax_amount(self) -> float:
        """Calculate tax amount."""
        return self.subtotal_after_discount * (self.tax_rate / 100)

    @rx.var
    def total(self) -> float:
        """Calculate total amount."""
        return self.subtotal_after_discount + self.tax_amount + self.shipping_cost

    @rx.event
    def on_load(self):
        """Initialize quotation with default values."""
        if not self.quote_date:
            today = datetime.now().strftime("%Y-%m-%d")
            self.quote_date = today
            # Valid for 30 days by default
            valid_date = datetime.now() + timedelta(days=30)
            self.valid_until = valid_date.strftime("%Y-%m-%d")

        # Add a default item if empty
        if not self.items:
            self.items = [
                QuotationItem(
                    id=str(uuid.uuid4()),
                    description="Servicio de logística",
                    quantity=1,
                    unit_price=0.0,
                    discount=0.0,
                    amount=0.0,
                    notes="",
                )
            ]

    @rx.event
    def set_field(self, field: str, value: str):
        """Update a single field."""
        # Handle numeric conversions
        if field in ["tax_rate", "shipping_cost", "discount_global"]:
            try:
                setattr(self, field, float(value) if value else 0.0)
            except ValueError:
                logging.warning(f"Invalid numeric value for {field}: {value}")
        else:
            setattr(self, field, value)

    @rx.event
    def add_item(self):
        """Add a new item to the quotation."""
        self.items.append(
            QuotationItem(
                id=str(uuid.uuid4()),
                description="Nuevo servicio",
                quantity=1,
                unit_price=0.0,
                discount=0.0,
                amount=0.0,
                notes="",
            )
        )

    @rx.event
    def remove_item(self, idx: int):
        """Remove an item from the quotation."""
        if 0 <= idx < len(self.items):
            self.items.pop(idx)
            # Create a new list to ensure Reflex detects the change correctly
            self.items = list(self.items)

    @rx.event
    def update_item(self, idx: int, field: str, value: str):
        """Update a specific field of an item."""
        if 0 <= idx < len(self.items):
            item = self.items[idx]

            # Handle different field types
            if field == "description" or field == "notes":
                setattr(item, field, value)
            elif field == "quantity":
                try:
                    item.quantity = int(value) if value else 1
                except ValueError:
                    item.quantity = 1
            elif field in ["unit_price", "discount"]:
                try:
                    setattr(item, field, float(value) if value else 0.0)
                except ValueError:
                    setattr(item, field, 0.0)

            # Recalculate amount
            item.amount = (item.quantity * item.unit_price) - item.discount

            # Trigger reactivity
            self.items = list(self.items)

    @rx.event
    def copy_to_clipboard(self):
        """Generate plain text version and copy to clipboard."""
        text_content = self._generate_text_content()

        # Escape backticks and newlines for JavaScript
        escaped_text = text_content.replace("`", "\\`").replace("\n", "\\n")

        return [
            rx.call_script(
                f"""
                navigator.clipboard.writeText(`{escaped_text}`).then(
                    () => console.log('Copied to clipboard'),
                    (err) => console.error('Failed to copy: ', err)
                );
                """
            ),
            rx.toast.success("Cotización copiada al portapapeles!"),
        ]

    def _generate_text_content(self) -> str:
        """Generate plain text version of quotation for email."""
        lines = [
            "=" * 60,
            f"COTIZACIÓN #{self.quote_number}",
            "=" * 60,
            f"Fecha: {self.quote_date}",
            f"Válida hasta: {self.valid_until}",
            "",
            "CLIENTE:",
            f"Nombre: {self.client_name}",
        ]

        if self.client_company:
            lines.append(f"Empresa: {self.client_company}")
        if self.client_address:
            lines.append(f"Dirección: {self.client_address}")
        if self.client_email:
            lines.append(f"Email: {self.client_email}")
        if self.client_phone:
            lines.append(f"Teléfono: {self.client_phone}")

        lines.extend(
            [
                "",
                "DETALLES:",
                "-" * 60,
            ]
        )

        for i, item in enumerate(self.items, 1):
            lines.append(f"{i}. {item.description}")
            lines.append(
                f"   Cantidad: {item.quantity} x ${item.unit_price:.2f} = ${item.amount:.2f}"
            )
            if item.discount > 0:
                lines.append(f"   Descuento: -${item.discount:.2f}")
            if item.notes:
                lines.append(f"   Nota: {item.notes}")

        lines.extend(["-" * 60, f"Subtotal: ${self.subtotal:.2f}"])

        if self.discount_global > 0:
            lines.append(f"Descuento Global: -${self.discount_global:.2f}")

        if self.tax_rate > 0:
            lines.append(f"Impuestos ({self.tax_rate}%): ${self.tax_amount:.2f}")

        if self.shipping_cost > 0:
            lines.append(f"Envío: ${self.shipping_cost:.2f}")

        lines.extend(["", f"TOTAL: ${self.total:.2f}", "=" * 60])

        if self.notes:
            lines.extend(["", "NOTAS:", self.notes])

        if self.payment_terms:
            lines.extend(["", "TÉRMINOS DE PAGO:", self.payment_terms])

        if self.terms_conditions:
            lines.extend(["", "TÉRMINOS Y CONDICIONES:", self.terms_conditions])

        lines.extend(
            [
                "",
                "Gracias por su confianza.",
                f"{self.company_name}",
                f"{self.company_phone}",
            ]
        )

        return "\n".join(lines)

    @rx.event
    async def export_pdf(self):
        """Generate and download PDF."""
        self.is_loading = True

        filename = f"Cotizacion_{self.quote_number}_{uuid.uuid4().hex[:6]}.pdf"
        upload_dir = Path(".web/public")
        upload_dir.mkdir(parents=True, exist_ok=True)
        file_path = upload_dir / filename

        try:
            # Create document
            doc = SimpleDocTemplate(
                str(file_path),
                pagesize=letter,
                rightMargin=40,
                leftMargin=40,
                topMargin=40,
                bottomMargin=40,
            )

            elements = []
            styles = getSampleStyleSheet()

            # Header section with logo and quotation info
            header_data = []
            logo_path = Path(".web/public") / self.logo_url.lstrip("/")

            if logo_path.exists():
                logo = Image(str(logo_path), width=0.8 * inch, height=0.8 * inch)
                company_info = Paragraph(
                    f"<b>{self.company_name}</b><br/>{self.company_address}<br/>{self.company_phone}",
                    styles["Normal"],
                )
                quote_info = Paragraph(
                    f"<b style='font-size:20; color:purple'>COTIZACIÓN</b><br/><b>No. {self.quote_number}</b><br/>Fecha: {self.quote_date}<br/>Válida hasta: {self.valid_until}",
                    styles["Normal"],
                )
                header_data.append([logo, company_info, quote_info])
            else:
                company_info = Paragraph(
                    f"<b>{self.company_name}</b><br/>{self.company_address}<br/>{self.company_phone}",
                    styles["Normal"],
                )
                quote_info = Paragraph(
                    f"<b style='font-size:20; color:purple'>COTIZACIÓN</b><br/><b>No. {self.quote_number}</b><br/>Fecha: {self.quote_date}<br/>Válida hasta: {self.valid_until}",
                    styles["Normal"],
                )
                header_data.append([company_info, quote_info])

            header_table = Table(
                header_data, colWidths=[1.5 * inch, 2.5 * inch, 2.5 * inch]
            )
            header_table.setStyle(
                TableStyle(
                    [
                        ("ALIGN", (0, 0), (0, 0), "LEFT"),
                        ("ALIGN", (1, 0), (1, 0), "LEFT"),
                        ("ALIGN", (2, 0), (2, 0), "RIGHT"),
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ]
                )
            )
            elements.append(header_table)
            elements.append(Spacer(1, 20))

            # Client section
            client_text = f"<b>PARA:</b><br/><b>{self.client_name}</b><br/>"
            if self.client_company:
                client_text += f"{self.client_company}<br/>"
            if self.client_address:
                client_text += f"{self.client_address}<br/>"
            if self.client_email:
                client_text += f"Email: {self.client_email}<br/>"
            if self.client_phone:
                client_text += f"Teléfono: {self.client_phone}<br/>"

            client_para = Paragraph(client_text, styles["Normal"])
            elements.append(client_para)
            elements.append(Spacer(1, 20))

            # Items table
            items_data = [
                ["DESCRIPCIÓN", "CANT.", "PRECIO", "DESC.", "TOTAL"]
            ]

            for item in self.items:
                discount_display = f"-${item.discount:.2f}" if item.discount > 0 else "-"
                items_data.append(
                    [
                        item.description,
                        str(item.quantity),
                        f"${item.unit_price:.2f}",
                        discount_display,
                        f"${item.amount:.2f}",
                    ]
                )

            items_table = Table(
                items_data, colWidths=[3 * inch, 0.6 * inch, 0.8 * inch, 0.8 * inch, 1 * inch]
            )
            items_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.Color(0.9, 0.9, 0.9)),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                        ("ALIGN", (0, 0), (0, -1), "LEFT"),
                        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, 0), 9),
                        ("FONTSIZE", (0, 1), (-1, -1), 9),
                        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                        ("TOPPADDING", (0, 1), (-1, -1), 8),
                        ("BOTTOMPADDING", (0, 1), (-1, -1), 8),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ]
                )
            )
            elements.append(items_table)
            elements.append(Spacer(1, 20))

            # Totals section
            totals_data = [
                ["Subtotal:", f"${self.subtotal:.2f}"],
            ]

            if self.discount_global > 0:
                totals_data.append(["Descuento:", f"-${self.discount_global:.2f}"])

            if self.tax_rate > 0:
                totals_data.append(
                    [f"Impuestos ({self.tax_rate}%):", f"${self.tax_amount:.2f}"]
                )

            if self.shipping_cost > 0:
                totals_data.append(["Envío:", f"${self.shipping_cost:.2f}"])

            totals_data.append(["TOTAL:", f"${self.total:.2f}"])

            totals_table = Table(totals_data, colWidths=[4.6 * inch, 1.6 * inch])
            totals_table.setStyle(
                TableStyle(
                    [
                        ("ALIGN", (0, 0), (0, -1), "RIGHT"),
                        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                        ("FONTSIZE", (0, -1), (-1, -1), 12),
                        ("TOPPADDING", (0, -1), (-1, -1), 10),
                        ("TEXTCOLOR", (0, -1), (-1, -1), colors.Color(0.5, 0, 0.5)),
                    ]
                )
            )
            elements.append(totals_table)
            elements.append(Spacer(1, 20))

            # Additional sections
            if self.notes:
                notes_para = Paragraph(
                    f"<b>NOTAS:</b><br/>{self.notes}", styles["Normal"]
                )
                elements.append(notes_para)
                elements.append(Spacer(1, 12))

            if self.payment_terms:
                payment_para = Paragraph(
                    f"<b>TÉRMINOS DE PAGO:</b><br/>{self.payment_terms}",
                    styles["Normal"],
                )
                elements.append(payment_para)
                elements.append(Spacer(1, 12))

            if self.terms_conditions:
                terms_para = Paragraph(
                    f"<b>TÉRMINOS Y CONDICIONES:</b><br/>{self.terms_conditions}",
                    styles["Normal"],
                )
                elements.append(terms_para)

            # Build PDF
            doc.build(elements)
            self.is_loading = False

            # Read file data and pass directly to download
            with open(file_path, "rb") as f:
                pdf_data = f.read()

            # Clean up temporary file
            try:
                file_path.unlink()
            except Exception as e:
                logging.warning(f"Could not delete temporary file: {e}")

            return rx.download(data=pdf_data, filename=filename)

        except Exception as e:
            self.is_loading = False
            logging.exception(f"PDF Generation Error: {e}")
            return rx.toast.error(f"Error generando PDF: {str(e)}")

    @rx.event
    async def export_excel(self):
        """Generate and download Excel file."""
        self.is_loading = True

        filename = f"Cotizacion_{self.quote_number}_{uuid.uuid4().hex[:6]}.xlsx"
        upload_dir = Path(".web/public")
        upload_dir.mkdir(parents=True, exist_ok=True)
        file_path = upload_dir / filename

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Cotización"

            # Styles
            title_font = Font(bold=True, size=16, color="800080")
            header_font = Font(bold=True, size=11)
            bold_font = Font(bold=True)
            gray_fill = PatternFill(
                start_color="EEEEEE", end_color="EEEEEE", fill_type="solid"
            )
            purple_fill = PatternFill(
                start_color="E6E6FA", end_color="E6E6FA", fill_type="solid"
            )

            # Title
            ws["A1"] = "COTIZACIÓN"
            ws["A1"].font = title_font
            ws.merge_cells("A1:E1")

            # Company info
            ws["A3"] = self.company_name
            ws["A3"].font = bold_font
            ws["A4"] = self.company_address
            ws["A5"] = self.company_phone

            # Quotation info
            ws["D3"] = "Número:"
            ws["E3"] = self.quote_number
            ws["D4"] = "Fecha:"
            ws["E4"] = self.quote_date
            ws["D5"] = "Válida hasta:"
            ws["E5"] = self.valid_until
            ws["D3"].font = bold_font
            ws["D4"].font = bold_font
            ws["D5"].font = bold_font

            # Client section
            row = 7
            ws[f"A{row}"] = "CLIENTE:"
            ws[f"A{row}"].font = bold_font
            row += 1
            ws[f"A{row}"] = self.client_name
            ws[f"A{row}"].font = bold_font

            if self.client_company:
                row += 1
                ws[f"A{row}"] = self.client_company

            if self.client_address:
                row += 1
                ws[f"A{row}"] = self.client_address

            if self.client_email:
                row += 1
                ws[f"A{row}"] = f"Email: {self.client_email}"

            if self.client_phone:
                row += 1
                ws[f"A{row}"] = f"Teléfono: {self.client_phone}"

            # Items table
            row += 2
            headers = ["DESCRIPCIÓN", "CANTIDAD", "PRECIO", "DESCUENTO", "TOTAL"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = gray_fill

            row += 1
            for item in self.items:
                ws.cell(row=row, column=1, value=item.description)
                ws.cell(row=row, column=2, value=item.quantity)
                ws.cell(row=row, column=3, value=item.unit_price)
                ws.cell(row=row, column=4, value=item.discount)
                ws.cell(row=row, column=5, value=item.amount)
                row += 1

            # Totals
            row += 1
            ws.cell(row=row, column=4, value="Subtotal:").font = bold_font
            ws.cell(row=row, column=5, value=self.subtotal)

            if self.discount_global > 0:
                row += 1
                ws.cell(row=row, column=4, value="Descuento:").font = bold_font
                ws.cell(row=row, column=5, value=-self.discount_global)

            if self.tax_rate > 0:
                row += 1
                ws.cell(
                    row=row, column=4, value=f"Impuestos ({self.tax_rate}%):"
                ).font = bold_font
                ws.cell(row=row, column=5, value=self.tax_amount)

            if self.shipping_cost > 0:
                row += 1
                ws.cell(row=row, column=4, value="Envío:").font = bold_font
                ws.cell(row=row, column=5, value=self.shipping_cost)

            row += 1
            total_cell_label = ws.cell(row=row, column=4, value="TOTAL:")
            total_cell_label.font = Font(bold=True, size=12, color="800080")
            total_cell_value = ws.cell(row=row, column=5, value=self.total)
            total_cell_value.font = Font(bold=True, size=12, color="800080")
            total_cell_value.fill = purple_fill

            # Column widths
            ws.column_dimensions["A"].width = 40
            ws.column_dimensions["B"].width = 12
            ws.column_dimensions["C"].width = 12
            ws.column_dimensions["D"].width = 18
            ws.column_dimensions["E"].width = 15

            # Save
            wb.save(file_path)
            self.is_loading = False

            # Read file data and pass directly to download
            with open(file_path, "rb") as f:
                excel_data = f.read()

            # Clean up temporary file
            try:
                file_path.unlink()
            except Exception as e:
                logging.warning(f"Could not delete temporary file: {e}")

            return rx.download(data=excel_data, filename=filename)

        except Exception as e:
            self.is_loading = False
            logging.exception(f"Excel Generation Error: {e}")
            return rx.toast.error(f"Error generando Excel: {str(e)}")
