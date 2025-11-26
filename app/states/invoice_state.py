import reflex as rx
from typing import Any
from datetime import datetime
from pathlib import Path
import uuid
import logging
from pydantic import BaseModel
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


class InvoiceItem(BaseModel):
    id: str
    code: str = ""           # Código/SKU
    description: str
    quantity: int
    unit_price: float
    discount: float = 0.0    # Descuento unitario
    amount: float
    tax_rate: float = 0.0   # Tasa de impuesto específica por item


class InvoiceState(rx.State):
    """State for the Invoice document."""

    is_loading: bool = False
    from_name: str = "Nosglobal Logistic"
    from_address: str = "Av. Principal 1000, Torre A, Piso 5"
    from_details: str = "Caracas, Distrito Capital, 1010"
    from_email: str = "info@nosglobal.com"
    from_phone: str = "+58 424-4966616"
    to_name: str = ""
    to_company: str = ""
    to_address: str = ""
    to_details: str = ""
    # Información fiscal
    from_tax_id: str = "J-123456789"  # RIF/Cédula del emisor
    to_tax_id: str = ""    # RIF/Cédula del cliente

    # Información de pago
    payment_method: str = ""
    bank_account: str = ""
    bank_name: str = ""

    # Información adicional
    terms_conditions: str = ""
    notes: str = ""
    authorized_by: str = ""
    logo_url: str = "/nosglobal-logo.png"

    invoice_number: str = ""
    invoice_date: str = ""
    due_date: str = ""
    items: list[InvoiceItem] = []
    tax_rate: float = 0.0

    @rx.event
    def on_load(self):
        """Initialize dates on client load to avoid hydration mismatch."""
        if not self.invoice_date:
            today = datetime.now().strftime("%Y-%m-%d")
            self.invoice_date = today
            self.due_date = today

    @rx.var
    def subtotal(self) -> float:
        return sum([item.amount for item in self.items])

    @rx.var
    def tax_amount(self) -> float:
        return self.subtotal * (self.tax_rate / 100)

    @rx.var
    def total(self) -> float:
        return self.subtotal + self.tax_amount

    @rx.event
    def set_field(self, field: str, value: str):
        if hasattr(self, field):
            if field == "tax_rate":
                try:
                    self.tax_rate = float(value)
                except ValueError as e:
                    logging.exception(f"Error converting tax_rate to float: {e}")
                    self.tax_rate = 0.0
            else:
                setattr(self, field, value)

    @rx.event
    def add_item(self):
        self.items.append(
            InvoiceItem(
                id=str(uuid.uuid4()),
                code="",
                description="Nuevo Item",
                quantity=1,
                unit_price=0.0,
                discount=0.0,
                amount=0.0,
                tax_rate=0.0,
            )
        )

    @rx.event
    def remove_item(self, idx: int):
        if 0 <= idx < len(self.items):
            self.items.pop(idx)
            # Create a new list to ensure Reflex detects the change correctly
            self.items = list(self.items)

    @rx.event
    def update_item(self, idx: int, field: str, value: Any):
        if 0 <= idx < len(self.items):
            item = self.items[idx]
            if field in ["quantity", "unit_price", "discount"]:
                try:
                    if value == "" or value is None:
                        val = 0.0
                    else:
                        val = float(value)

                    if field == "quantity":
                        item.quantity = int(val)
                    elif field == "unit_price":
                        item.unit_price = val
                    elif field == "discount":
                        item.discount = val
                    # Recalculate amount with discount
                    item.amount = item.quantity * (item.unit_price - item.discount)
                except ValueError as e:
                    logging.warning(f"Invalid value for field {field}: {value}")
            elif field in ["description", "code", "tax_rate"]:
                if field == "tax_rate":
                    try:
                        if value == "" or value is None:
                            item.tax_rate = 0.0
                        else:
                            item.tax_rate = float(value)
                    except ValueError as e:
                        logging.warning(f"Invalid value for field {field}: {value}")
                else:
                    setattr(item, field, str(value))
            
            # Create a new list to ensure Reflex detects the change correctly
            self.items = list(self.items)

    @rx.event
    async def export_pdf(self):
        self.is_loading = True
        filename = f"NotaDeEntrega_{self.invoice_number}_{uuid.uuid4().hex[:6]}.pdf"
        upload_dir = Path(".web/public")
        upload_dir.mkdir(parents=True, exist_ok=True)
        file_path = upload_dir / filename
        try:
            doc = SimpleDocTemplate(
                str(file_path),
                pagesize=letter,
                rightMargin=40,
                leftMargin=40,
                topMargin=40,
                bottomMargin=40,
            )
            elements = []
            styles = getSampleStyleSheet()
            header_data = [
                [
                    Paragraph(
                        f"<b>{self.from_name}</b><br/>{self.from_address}<br/>{self.from_details}<br/>RIF/Cédula: {self.from_tax_id}<br/>{self.from_email}<br/>{self.from_phone}",
                        styles["Normal"],
                    ),
                    Paragraph(
                        f"<font size=16><b>NOTA DE ENTREGA</b></font><br/><br/><b>No:</b> {self.invoice_number}<br/><b>Fecha:</b> {self.invoice_date}<br/><b>Vence:</b> {self.due_date}",
                        styles["Normal"],
                    ),
                ]
            ]
            t_header = Table(header_data, colWidths=[4 * inch, 3 * inch])
            t_header.setStyle(
                TableStyle(
                    [
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                        ("ALIGN", (1, 0), (1, 0), "RIGHT"),
                    ]
                )
            )
            elements.append(t_header)
            elements.append(Spacer(1, 30))
            elements.append(Paragraph("<b>ENTREGAR A:</b>", styles["Heading4"]))
            elements.append(
                Paragraph(
                    f"{self.to_name}<br/>{self.to_company}<br/>{self.to_address}<br/>{self.to_details}<br/>RIF/Cédula: {self.to_tax_id}",
                    styles["Normal"],
                )
            )
            elements.append(Spacer(1, 30))
            data = [["CÓDIGO", "DESCRIPCIÓN", "CANT.", "PRECIO", "DESC.", "TOTAL"]]
            for item in self.items:
                description_text = f"{item.code} - {item.description}" if item.code else item.description
                discount_text = f"${item.discount:.2f}" if item.discount > 0 else "-"
                data.append(
                    [
                        Paragraph(item.code if item.code else "-", styles["Normal"]),
                        Paragraph(description_text, styles["Normal"]),
                        str(item.quantity),
                        f"${item.unit_price:,.2f}",
                        discount_text,
                        f"${item.amount:,.2f}",
                    ]
                )
            t_items = Table(
                data, colWidths=[0.8 * inch, 2.5 * inch, 0.8 * inch, 1 * inch, 0.8 * inch, 1.2 * inch]
            )
            t_items.setStyle(
                TableStyle(
                    [
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, 0), 9),
                        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                        ("LINEBELOW", (0, 0), (-1, 0), 1, colors.black),
                        ("ALIGN", (2, 0), (-1, -1), "RIGHT"),
                        ("ALIGN", (0, 0), (0, -1), "CENTER"),
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.gray),
                        ("TEXTCOLOR", (4, 1), (4, -1), colors.red),  # Discount column in red
                    ]
                )
            )
            elements.append(t_items)
            elements.append(Spacer(1, 20))
            totals_data = [
                ["Subtotal:", f"${self.subtotal:,.2f}"],
                [f"Impuestos ({self.tax_rate}%):", f"${self.tax_amount:,.2f}"],
                ["Total:", f"${self.total:,.2f}"],
            ]
            t_totals = Table(totals_data, colWidths=[4.75 * inch, 1.25 * inch])
            t_totals.setStyle(
                TableStyle(
                    [
                        ("ALIGN", (0, 0), (-1, -1), "RIGHT"),
                        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                        ("FONTSIZE", (0, -1), (-1, -1), 12),
                        ("LINEABOVE", (0, -1), (-1, -1), 1, colors.black),
                        ("TOPPADDING", (0, -1), (-1, -1), 10),
                    ]
                )
            )
            elements.append(t_totals)
            elements.append(Spacer(1, 30))
            
            # Payment Information Section
            if self.payment_method:
                elements.append(Paragraph("<b>INFORMACIÓN DE PAGO:</b>", styles["Heading4"]))
                payment_info = f"Método: {self.payment_method}<br/>"
                if self.bank_name:
                    payment_info += f"Banco: {self.bank_name}<br/>"
                if self.bank_account:
                    payment_info += f"Cuenta: {self.bank_account}"
                elements.append(Paragraph(payment_info, styles["Normal"]))
                elements.append(Spacer(1, 20))
            
            # Terms and Conditions Section
            if self.terms_conditions:
                elements.append(Paragraph("<b>TÉRMINOS Y CONDICIONES:</b>", styles["Heading4"]))
                elements.append(Paragraph(self.terms_conditions, styles["Normal"]))
                elements.append(Spacer(1, 20))
            
            # Notes Section
            if self.notes:
                elements.append(Paragraph("<b>NOTAS:</b>", styles["Heading4"]))
                elements.append(Paragraph(self.notes, styles["Normal"]))
                elements.append(Spacer(1, 20))
            
            # Authorization Section
            if self.authorized_by:
                elements.append(Paragraph("<b>AUTORIZACIÓN:</b>", styles["Heading4"]))
                auth_data = [
                    ["", ""],
                    [f"Autorizado por: {self.authorized_by}", "Firma:"],
                    ["", ""],
                ]
                t_auth = Table(auth_data, colWidths=[3 * inch, 3 * inch])
                t_auth.setStyle(
                    TableStyle(
                        [
                            ("ALIGN", (0, 1), (0, 1), "LEFT"),
                            ("ALIGN", (1, 1), (1, 1), "CENTER"),
                            ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
                            ("LINEBELOW", (1, 1), (1, 1), 1, colors.black),
                            ("BOTTOMPADDING", (1, 1), (1, 1), 20),
                        ]
                    )
                )
                elements.append(t_auth)
            
            doc.build(elements)
            self.is_loading = False
            if not file_path.exists():
                raise FileNotFoundError(f"PDF file was not created: {file_path}")

            # Read file data and pass directly to download
            with open(file_path, "rb") as f:
                pdf_data = f.read()

            # Clean up temporary file
            try:
                file_path.unlink()
            except Exception as e:
                logging.warning(f"Could not delete temporary file: {e}")

            return rx.download(data=pdf_data, filename=filename)
        except Exception as e:
            self.is_loading = False
            logging.exception(f"PDF Generation Error: {e}")
            return rx.toast.error(f"Error generating PDF: {str(e)}")

    @rx.event
    async def export_excel(self):
        self.is_loading = True
        filename = f"NotaDeEntrega_{self.invoice_number}_{uuid.uuid4().hex[:6]}.xlsx"
        upload_dir = Path(".web/public")
        upload_dir.mkdir(parents=True, exist_ok=True)
        file_path = upload_dir / filename
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Nota de entrega"
            title_font = Font(bold=True, size=16)
            header_font = Font(bold=True)
            gray_fill = PatternFill(
                start_color="EEEEEE", end_color="EEEEEE", fill_type="solid"
            )
            ws["A1"] = self.from_name
            ws["A1"].font = title_font
            ws["A2"] = self.from_address
            ws["A3"] = self.from_details
            ws["A4"] = self.from_email
            ws["A5"] = f"RIF/Cédula: {self.from_tax_id}"
            ws["A6"] = self.from_phone
            ws["E1"] = "NOTA DE ENTREGA"
            ws["E1"].font = title_font
            ws["E2"] = f"No: {self.invoice_number}"
            ws["E3"] = f"Fecha: {self.invoice_date}"
            ws["E4"] = f"Vence: {self.due_date}"
            ws["A8"] = "ENTREGAR A:"
            ws["A8"].font = header_font
            ws["A9"] = self.to_name
            ws["A10"] = self.to_company
            ws["A11"] = self.to_address
            ws["A12"] = self.to_details
            ws["A13"] = f"RIF/Cédula: {self.to_tax_id}"
            row = 15
            headers = ["Código", "Descripción", "Cantidad", "Precio Unitario", "Descuento", "Total"]
            for col, text in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=text)
                cell.font = header_font
                cell.fill = gray_fill
            row += 1
            for item in self.items:
                ws.cell(row=row, column=1, value=item.code if item.code else "-")
                description = f"{item.code} - {item.description}" if item.code else item.description
                ws.cell(row=row, column=2, value=description)
                ws.cell(row=row, column=3, value=item.quantity)
                ws.cell(row=row, column=4, value=item.unit_price)
                ws.cell(row=row, column=5, value=item.discount if item.discount > 0 else 0)
                ws.cell(row=row, column=6, value=item.amount)
                row += 1
            row += 2
            ws.cell(row=row, column=4, value="Subtotal:").font = header_font
            ws.cell(row=row, column=5, value=self.subtotal)
            ws.cell(row=row, column=6, value=self.subtotal)
            row += 1
            ws.cell(
                row=row, column=4, value=f"Impuestos ({self.tax_rate}%):"
            ).font = header_font
            ws.cell(row=row, column=5, value=self.tax_amount)
            ws.cell(row=row, column=6, value=self.tax_amount)
            row += 1
            ws.cell(row=row, column=4, value="TOTAL:").font = header_font
            ws.cell(row=row, column=5, value=self.total)
            ws.cell(row=row, column=6, value=self.total).font = Font(bold=True)
            
            # Add payment information if available
            if self.payment_method:
                row += 2
                ws.cell(row=row, column=1, value="INFORMACIÓN DE PAGO:").font = header_font
                row += 1
                ws.cell(row=row, column=1, value=f"Método: {self.payment_method}")
                if self.bank_name:
                    row += 1
                    ws.cell(row=row, column=1, value=f"Banco: {self.bank_name}")
                if self.bank_account:
                    row += 1
                    ws.cell(row=row, column=1, value=f"Cuenta: {self.bank_account}")
            
            # Add terms and conditions if available
            if self.terms_conditions:
                row += 2
                ws.cell(row=row, column=1, value="TÉRMINOS Y CONDICIONES:").font = header_font
                row += 1
                ws.cell(row=row, column=1, value=self.terms_conditions)
            
            # Add notes if available
            if self.notes:
                row += 2
                ws.cell(row=row, column=1, value="NOTAS:").font = header_font
                row += 1
                ws.cell(row=row, column=1, value=self.notes)
            
            # Add authorization if available
            if self.authorized_by:
                row += 2
                ws.cell(row=row, column=1, value="AUTORIZACIÓN:").font = header_font
                row += 1
                ws.cell(row=row, column=1, value=f"Autorizado por: {self.authorized_by}")
                row += 1
                ws.cell(row=row, column=1, value="Firma:")
            
            ws.column_dimensions["A"].width = 40
            ws.column_dimensions["B"].width = 30
            ws.column_dimensions["C"].width = 12
            ws.column_dimensions["D"].width = 15
            ws.column_dimensions["E"].width = 15
            ws.column_dimensions["F"].width = 15
            wb.save(file_path)
            self.is_loading = False

            # Read file data and pass directly to download
            with open(file_path, "rb") as f:
                excel_data = f.read()

            # Clean up temporary file
            try:
                file_path.unlink()
            except Exception as e:
                logging.warning(f"Could not delete temporary file: {e}")

            return rx.download(data=excel_data, filename=filename)
        except Exception as e:
            self.is_loading = False
            logging.exception(f"Excel Generation Error: {e}")
            return rx.toast.error(f"Error generating Excel: {str(e)}")